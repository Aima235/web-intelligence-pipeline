from pathlib import Path
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class DatasetGenerator:

    def __init__(self):

        self.json_dir = Path("output/json")

        self.output_dir = Path("output/dataset")
        self.output_dir.mkdir(parents=True, exist_ok=True)

        self.output_file = self.output_dir / "market_gap_dataset.xlsx"

    def _join(self, value):

        if isinstance(value, list):
            return ", ".join(str(v) for v in value)

        return value if value else ""

    def generate(self):

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Market Gap Dataset"

        headers = [
            "Company Name",
            "Website",
            "Industry",
            "Company Summary",
            "Services",
            "Technologies",
            "Business Gaps",
            "AI Opportunities",
            "Match Score",
            "Confidence",
            "Matched Skills",
            "Missing Skills",
            "Strengths",
            "Improvement Areas",
            "Recommended Services",
            "Lead Priority",
            "Lead Score",
            "Contact Page",
            "Candidate Name",
            "Desired Role",
            "No. of Services",
            "No. of Technologies",
            "No. of Business Gaps",
            "No. of Matched Skills",
            "No. of Missing Skills",
            "No. of Recommended Services"
        ]

        sheet.append(headers)

        total = 0

        for file in sorted(self.json_dir.glob("*.json")):

            try:

                with open(file, "r", encoding="utf-8") as f:
                    data = json.load(f)

                company = data.get("company", {})
                match = data.get("candidate_match", {})
                lead = data.get("lead_generation", {})
                profile = data.get("candidate_profile", {})

                row = [

                    company.get("name", ""),
                    company.get("website", ""),
                    company.get("industry", ""),
                    company.get("summary", ""),

                    self._join(data.get("services", [])),
                    self._join(data.get("technologies", [])),
                    self._join(data.get("business_gaps", [])),
                    self._join(data.get("ai_opportunities", [])),

                    match.get("score", ""),
                    match.get("confidence", ""),
                    self._join(match.get("matched_skills", [])),
                    self._join(match.get("missing_skills", [])),
                    self._join(match.get("strengths", [])),
                    self._join(match.get("improvement_areas", [])),

                    self._join(data.get("recommended_services", [])),

                    lead.get("priority", ""),
                    lead.get("lead_score", ""),
                    lead.get("contact_page", ""),

                    profile.get("name", ""),
                    profile.get("desired_role", ""),

                    len(data.get("services", [])),
                    len(data.get("technologies", [])),
                    len(data.get("business_gaps", [])),
                    len(match.get("matched_skills", [])),
                    len(match.get("missing_skills", [])),
                    len(data.get("recommended_services", []))

                ]

                sheet.append(row)
                total += 1

            except Exception as e:
                print(f"Failed to process {file.name}: {e}")

        # -----------------------------
        # Formatting
        # -----------------------------

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="4F81BD"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for cell in sheet[1]:

            cell.fill = header_fill
            cell.font = header_font

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            cell.border = thin_border

        for row in sheet.iter_rows(min_row=2):

            for cell in row:

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

                cell.border = thin_border

        # Freeze first row
        sheet.freeze_panes = "A2"

        # Enable filters
        sheet.auto_filter.ref = sheet.dimensions

        # Auto-size columns
        for column_cells in sheet.columns:

            max_length = 0
            column = get_column_letter(column_cells[0].column)

            for cell in column_cells:

                try:
                    if cell.value:
                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )
                except Exception:
                    pass

            adjusted_width = min(max_length + 3, 50)

            sheet.column_dimensions[column].width = adjusted_width

        workbook.save(self.output_file)

        print("\n==========================================")
        print("DATASET CREATED SUCCESSFULLY")
        print("==========================================")
        print(f"Companies Processed : {total}")
        print(f"Saved To            : {self.output_file}")
        print("==========================================")